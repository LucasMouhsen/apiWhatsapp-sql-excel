# pip install pandas pyodbc sqlalchemy
import unicodedata 
from unidecode import unidecode
import pandas as pd
import email
import email.mime.text
import email.mime.application
from email.mime.multipart import MIMEMultipart
import pyodbc
import sqlalchemy as sa
import smtplib
from email.mime.text import MIMEText
import re
from email.mime.application import MIMEApplication
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import Workbook
from urllib.parse import quote

# Conexión con base de datos
con_salud = sa.create_engine('mssql+pyodbc://user:password@ip/database?driver=SQL Server')


# Consulta a base de datos
consulta = f""

df = pd.read_sql(consulta, con_salud)  

# Generamos el mensaje base con las variables para despues reemplazar con los datos de la base de datos
mensaje_base= u'Buenos+dias+{nombre}+esto+es+un+mensaje+de+prueba+{especialidad}+{mediMedico}+{fechaTurno}.' 


# Normaliza el numero que viene por base
def parse_num(numero):
    if numero:
        numeros = re.sub(r'\D', '', numero)
        if len(numeros) <= 12:
            link = numeros
            return link
        else:
            numeros = numeros[-10:]
            link = "54"+numeros
            return link

# Creamos array vacio para agregar los mensajes
mensajes = []

# iteramos las filas de los datos obtenidos
for indice, fila in df.iterrows():
    # Declaramos las variables
    nombre = unidecode(fila['paci_Paciente']).strip()
    fechaTurno = fila['turn_FechaTurno'].strftime('%d/%m/%Y %H:%M')
    especialidad = unidecode(fila['nome_Descripcion']).strip()
    numero = parse_num(fila["tele_Numero"])
    mediMedico = unidecode(fila["medi_Medico"]).strip()
    # Le damos formato al mensaje
    mensaje = mensaje_base.format(nombre=nombre, fechaTurno=fechaTurno, especialidad=especialidad, mediMedico=mediMedico)
    # Reemplazamos los espacios por el signo "+"
    mensaje_mod = mensaje.replace(' ', '+')
    # Generamos el link con el mensaje y el numero parseado
    linklarge = f'https://wa.me/+{parse_num(numero)}?text={mensaje_mod}'
    # Agregamos el mensaje al array
    mensajes.append(linklarge)

# Agregamos los datos al dataframe
df_mensajes = pd.DataFrame(mensajes)
df_mensajes.columns = ["Mensaje"]
df_mensajes["pers_NumeroDocumento"] = df["pers_NumeroDocumento"]
df_mensajes["tele_Numero"] = df["tele_Numero"].apply(parse_num)
df_mensajes["nome_Descripcion"] = df["nome_Descripcion"]
df_mensajes["turn_Fecha"] = df["turn_Fecha"]
df_mensajes['turn_Codigo'] = df['turn_Codigo']
df_mensajes['paciCodigo'] = df['paciCodigo']

# Guardamos el archivo con extension xlsx
fechaArchivo = str(df["turn_Fecha"].values[0])
fechaArchivo = fechaArchivo.replace('/', '-')
nombre_archivo_xlsx = f'LINK-TURNOS-AUSENTISMO-{fechaArchivo}.xlsx'

wb = Workbook()
ws = wb.active

# añadimos encabezados 
for row in dataframe_to_rows(df_mensajes, index=False, header=True):
    ws.append(row)

# Hacemos que la primer columna sean todos links
ws.column_dimensions['A'].width = 50
ws['A1'].font = Font(underline="single", color="0563C1")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)):
    link = df_mensajes.loc[row_idx, "Mensaje"]
    link_cell = row[0]
    link_cell.hyperlink = link
    link_cell.value = "Link"
    link_cell.style = "Hyperlink"
    link_cell.font = Font(underline="single", color="0563C1")

# Guardamos el archivo xlsx
wb.save(nombre_archivo_xlsx)

# Configuración del servidor SMTP de Gmail y las credenciales de la cuenta
smtp_host = 'smtp.gmail.com'
smtp_port = 587
smtp_username = 'mail'
smtp_password = 'password'

# Configurar el objeto MIMEMultipart con el contenido del correo electrónico
mensaje = MIMEMultipart()
mensaje['Subject'] = 'asunto'
mensaje['From'] = smtp_username
mensaje['To'] = 'quien lo envia'

# Agregar el cuerpo del correo electrónico
cuerpo = MIMEText('cuerpo del mail')
mensaje.attach(cuerpo)

# Adjuntar el archivo Excel al correo electrónico
with open(nombre_archivo_xlsx, 'rb') as archivo:
    adjunto = MIMEApplication(archivo.read(), _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')

adjunto.add_header('Content-Disposition', 'attachment', filename=nombre_archivo_xlsx)
mensaje.attach(adjunto)

# Crear una conexión segura al servidor SMTP de Gmail
smtp_conn = smtplib.SMTP(smtp_host, smtp_port)
smtp_conn.starttls()
smtp_conn.login(smtp_username, smtp_password)

# Enviar el correo electrónico
remitentes = ['ejemplo@msm.gov.ar', 'prueba@msm.gov.ar']
smtp_conn.sendmail(smtp_username, remitentes, mensaje.as_string())


print('Correo enviado correctamente a', remitentes)

# Cerrar la conexión
smtp_conn.quit()